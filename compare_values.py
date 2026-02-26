"""Compare gdp_projections_v4.xlsx values with data.js values for each state, every year."""
import openpyxl, re

# Read Excel
wb = openpyxl.load_workbook('gdp_projections_v4.xlsx', data_only=True)
ws = wb['Year-Wise Projections']

# Row 4 has headers: None, State, 2026, 2027, ..., 2047
header_row = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
years = []
for c in range(3, ws.max_column + 1):
    v = ws.cell(4, c).value
    if v and str(v).isdigit():
        years.append((c, int(v)))

excel_data = {}
for r in range(5, ws.max_row + 1):
    state_name = ws.cell(r, 2).value
    if not state_name:
        continue
    excel_data[state_name] = {}
    for col, yr in years:
        val = ws.cell(r, col).value
        if val is not None:
            excel_data[state_name][yr] = float(val)

# Read data.js
with open('js/data.js', 'r', encoding='utf-8') as f:
    content = f.read()

# Extract each state's GDP anchors from data.js
pattern = r'name:\s*"([^"]+)",\s*code:\s*"([^"]+)",\s*popMillions:\s*([\d.]+),\s*gdpAnchors:\s*\{([^}]+)\}'
js_data = {}
for m in re.finditer(pattern, content):
    name = m.group(1)
    anchors_str = m.group(4)
    anchors = {}
    for a in re.finditer(r'(\d{4}):\s*([\d.]+)', anchors_str):
        anchors[int(a.group(1))] = float(a.group(2))
    js_data[name] = anchors

# Compare
print("=" * 100)
print("COMPARISON: Excel vs data.js  (only showing differences)")
print("=" * 100)

mismatches = 0
for state_name in excel_data:
    if state_name not in js_data:
        print(f"\n*** {state_name}: NOT FOUND in data.js ***")
        mismatches += 1
        continue
    
    for yr in sorted(excel_data[state_name]):
        excel_val = excel_data[state_name][yr]
        js_val = js_data[state_name].get(yr)
        if js_val is None:
            print(f"  {state_name} {yr}: Excel={excel_val}, data.js=MISSING")
            mismatches += 1
        elif abs(excel_val - js_val) > 0.05:
            print(f"  {state_name} {yr}: Excel={excel_val}, data.js={js_val}, diff={excel_val - js_val:.2f}")
            mismatches += 1

if mismatches == 0:
    print("\nAll 2026-2047 values MATCH exactly between Excel and data.js!")
else:
    print(f"\nFound {mismatches} mismatches")

# Also show what data.js has for 2024/2025 (which aren't in Excel)
print("\n" + "=" * 100)
print("data.js 2024 base values (NOT present in Excel — from MOSPI estimate):")
print("=" * 100)
for state_name in js_data:
    if 2024 in js_data[state_name]:
        print(f"  {state_name}: 2024 = ${js_data[state_name][2024]}B")
    if 2025 in js_data[state_name]:
        print(f"  {state_name}: 2025 = ${js_data[state_name][2025]}B")

# Show what the user currently sees (Feb 2026 live values)
print("\n" + "=" * 100)
print("LIVE VALUES (Feb 26, 2026 — ~15.6% through the year):")
print("What the user SEES vs what the Excel says for 2026:")
print("=" * 100)
import math
frac = 57 / 365.0  # Feb 26 = day 57 of 365
india_gdp_2025 = 4270
india_gdp_2026 = 4612
current_india_gdp = india_gdp_2025 + (india_gdp_2026 - india_gdp_2025) * frac
scale = current_india_gdp / india_gdp_2026
print(f"India live GDP: ${current_india_gdp:.1f}B (scale factor: {scale:.4f})")
print()

for state_name in excel_data:
    if state_name in js_data and 2026 in excel_data[state_name]:
        excel_2026 = excel_data[state_name][2026]
        # What the live display shows (scaled down since we're only in Feb)
        live_val = excel_2026 * scale
        print(f"  {state_name:<22} Excel 2026: ${excel_2026:>8.1f}B  |  Live display: ${live_val:>8.1f}B")
